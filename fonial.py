import argparse
import csv
import json
import logging
import os
from getpass import getpass
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup

class Employee:
    def __init__(self, name, number, extension, mac_address, fonial_number, fonial_device):
        self.name = name
        self.number = str(number)
        self.extension = extension
        self.mac_address = mac_address
        self.fonial_number = fonial_number
        self.fonial_device = fonial_device

    def __str__(self):
        return "{}, {} ({}), {} with fonial device {} and number {}".format(self.name, self.number, self.extension, self.mac_address, self.fonial_device, self.fonial_number)

class FonialDevice:
    def __init__(self, id, mac_address):
        self.id = id
        self.mac_address = mac_address

    def __str__(self):
        return "{}: {}".format(self.id, self.mac_address)

class FonialNumber:
    def __init__(self, id, number, type, state, cancelled, assigned):
        self.id = id
        self.number = str(number)
        self.extension = int(self.number[-3:])
        self.type = type
        self.state = state
        self.cancelled = cancelled
        self.assigned = assigned

    def __str__(self):
        return "{}: {} {} ({}) {}".format(self.id, self.type, self.number, "cancelled" if self.cancelled else self.state, "assigned" if self.assigned else "not assigned")

class Fonial(object):
    def __init__(self, args) -> None:
        self.args = args
        self.session = requests.session()
        self.login(self.args.user, self.args.password)
        logging.basicConfig(level=logging.DEBUG if self.args.debug else logging.INFO)


    def login(self, username, password):
        response = self.session.get("https://kundenkonto.fonial.de/login")
        soup = BeautifulSoup(response.content, 'html.parser')
        assert soup.title.string == "Login"
        csrf = soup.find("input", {"name": "_csrf_token"}).get("value")

        response = self.session.post("https://kundenkonto.fonial.de/login_check", data={
            "_csrf_token": csrf,
            "_username":  username,
            "_password": password
        })
        soup = BeautifulSoup(response.content, 'html.parser')
        assert soup.title.string == "fonial Kundenkonto"

    ## DEVICES

    def loadDevices(self):
        try:
            with open('devices.response', 'r') as f:
                logging.info("using cached devices from {}".format(os.path.realpath(f.name)))
                soup = BeautifulSoup(f.read(), 'html.parser')
        except IOError:
            logging.info("reload devices from fonial")
            response = self.session.post("https://kundenkonto.fonial.de/system/device/")
            soup = BeautifulSoup(response.content, 'html.parser')
            logging.debug(soup.prettify())
            with open('devices.response', 'w') as f:
                logging.debug("caching devices into {}".format(os.path.realpath(f.name)))
                f.write(response.text)

        devices = {}
        for tag in soup.select("#dataTable-ip tbody tr"):
            device = FonialDevice(tag.get("id"), tag.find_all("td")[2].string)
            devices[device.mac_address] = device
        logging.debug("found devices {}".format(devices))
        return devices

    def new_devices(self):
        with open(self.args.file, 'r', encoding='utf-8-sig', newline='') as src:
            reader = csv.reader(src, delimiter=';', quotechar='"')
            next(reader)  # skip header row

            numbers = self.loadNumbers()
            csrf = None

            for row in reader:
                print("creating {}".format(row))

                if not row[1]:
                    print("skip because no phone number found in CSV")
                    continue

                if not row[3]:
                    print("skip because no Mac Address found in CSV")
                    continue

                if not row[1] in numbers:
                    print("skip because phone number {} not found in fonial account".format(row[1]))
                    continue
                n = numbers[row[1]]

                csrf = self.new_device(row[0], row[3], n, csrf)

                # if not n.state:
                #     self.activateNumber(n)
                #
                # response = self.session.get("https://kundenkonto.fonial.de/system/device/ipdevice/new")
                # soup = BeautifulSoup(response.content, 'html.parser')
                # logging.debug(soup.prettify())
                # assert soup.form.get("name") == "fonial_frontend_device_ip_device"
                #
                # csrf = soup.find("input", id="fonial_frontend_device_ip_device__token").get("value")
                # assert csrf is not None
                # logging.info("found CSRF token {}".format(csrf))
                #
                # response = self.session.post("https://kundenkonto.fonial.de/system/device/create", data={
                #     "fonial_frontend_device_ip_device[model]": "13",            # Snom D715
                #     "fonial_frontend_device_ip_device[targetName]": row[0],     # <Name>
                #     "fonial_frontend_device_ip_device[mac]": row[3],            # <Mac-Address>
                #     "fonial_frontend_device_ip_device[provisioning_template]": "",
                #     "fonial_frontend_device_ip_device[keyextensions]": "",
                #     "fonial_frontend_device_ip_device[outboundnum]": n.id,      # the ID of the fonial phone number
                #     "fonial_frontend_device_ip_device[moh]": "4",
                #     "fonial_frontend_device_ip_device[internalext]": int(row[2]),   # the 3-digit extension 678
                #     "fonial_frontend_device_ip_device[automatic]": "1",
                #     "fonial_frontend_device_ip_device[hash]": "ip",
                #     "fonial_frontend_device_ip_device[targetType]": "IPDEVICE",
                #     "fonial_frontend_device_ip_device[account]": self.args.account,  # <fonial account id>,
                #     "fonial_frontend_device_ip_device[targetOtherPhone]": "",
                #     "fonial_frontend_device_ip_device[targetOtherModel]": "13",
                #     "fonial_frontend_device_ip_device[_token]": csrf            # the CSRF token
                # })
                # logging.debug(response.text)
                # assert response.status_code == 200
                # assert response.text == '["\\/system\\/device\\/#ip"]'

    def new_device(self, name, mac_address, number, csrf=None):
        if not number.state:
            self.activateNumber(number)

        if not csrf:
            response = self.session.get("https://kundenkonto.fonial.de/system/device/ipdevice/new")
            soup = BeautifulSoup(response.content, 'html.parser')
            logging.debug(soup.prettify())
            assert soup.form.get("name") == "fonial_frontend_device_ip_device"

            csrf = soup.find("input", id="fonial_frontend_device_ip_device__token").get("value")
            assert csrf is not None
            logging.debug("found CSRF token {}".format(csrf))

        response = self.session.post("https://kundenkonto.fonial.de/system/device/create", data={
            "fonial_frontend_device_ip_device[model]": "13",  # Snom D715
            "fonial_frontend_device_ip_device[targetName]": name,  # <Name>
            "fonial_frontend_device_ip_device[mac]": mac_address,  # <Mac-Address>
            "fonial_frontend_device_ip_device[provisioning_template]": "",
            "fonial_frontend_device_ip_device[keyextensions]": "",
            "fonial_frontend_device_ip_device[outboundnum]": number.id,  # the ID of the fonial phone number
            "fonial_frontend_device_ip_device[moh]": "4",
            "fonial_frontend_device_ip_device[internalext]": number.extension,  # the 3-digit extension - e.g. 678
            "fonial_frontend_device_ip_device[automatic]": "1",
            "fonial_frontend_device_ip_device[hash]": "ip",
            "fonial_frontend_device_ip_device[targetType]": "IPDEVICE",
            "fonial_frontend_device_ip_device[account]": self.args.account,  # <fonial account id>,
            "fonial_frontend_device_ip_device[targetOtherPhone]": "",
            "fonial_frontend_device_ip_device[targetOtherModel]": "13",
            "fonial_frontend_device_ip_device[_token]": csrf  # the CSRF token
        })
        logging.debug(response.text)
        assert response.status_code == 200
        assert response.text == '["\\/system\\/device\\/#ip"]'

        return csrf

    def update_device(self, device, user):
        response = self.session.get("https://kundenkonto.fonial.de/system/device/70784/ipdevice/edit")
        soup = BeautifulSoup(response.content, 'html.parser')
        logging.debug(soup.prettify())
        assert soup.form.get("name") == "fonial_frontend_device_ip_device"

    def delete_device(self, d: FonialDevice):
        logging.info("deleting device {}".format(d))
        response = self.session.post("https://kundenkonto.fonial.de/system/device/{}/delete".format(d.id))

    def bind_number_to_device(self, name, device: FonialDevice, number: FonialNumber, csrf=None):
        if not number.state:
            logging.info("{}: outbound number {} inactive".format(device.name, number.number))
            return

        logging.info("{}: fixing outbound number".format(device.id, number))
        if self.args.dry_run:
            return

        if not csrf:
            response = self.session.get("https://kundenkonto.fonial.de/system/device/{}/ipdevice/edit".format(device.id))
            soup = BeautifulSoup(response.content, 'html.parser')
            assert soup.form.get("name") == "fonial_frontend_device_ip_device"
            csrf = soup.find("input", id="fonial_frontend_device_ip_device__token").get("value")
            assert csrf is not None
            logging.debug("found CSRF token {}".format(csrf))

        response = self.session.post(
            "https://kundenkonto.fonial.de/system/device/{}/update".format(device.id), data={
                "fonial_frontend_device_ip_device[targetName]": name,  # <Name>
                "fonial_frontend_device_ip_device[mac]": device.mac_address,  # <Mac-Address>
                "fonial_frontend_device_ip_device[provisioning_template]": "",
                "fonial_frontend_device_ip_device[keyextensions]": "",
                "fonial_frontend_device_ip_device[outboundnum]": number.id,
                # the ID of the fonial phone number
                "fonial_frontend_device_ip_device[moh]": "4",
                "fonial_frontend_device_ip_device[internalext]": number.extension,  # the 3-digit extension 678
                "fonial_frontend_device_ip_device[automatic]": "1",
                "fonial_frontend_device_ip_device[hash]": "ip",
                "fonial_frontend_device_ip_device[targetType]": "IPDEVICE",
                "fonial_frontend_device_ip_device[account]": self.args.account,  # <fonial account id>
                "fonial_frontend_device_ip_device[targetOtherPhone]": "",
                "fonial_frontend_device_ip_device[targetOtherModel]": "13",
                "fonial_frontend_device_ip_device[_token]": csrf  # the CSRF token
            })


    ## NUMBERS

    def loadNumbers(self):
        response = self.session.post("https://kundenkonto.fonial.de/system/number/json", data={
            "draw": "1",
            "columns[0][data]": "number",
            "columns[0][name]": "",
            "columns[0][searchable]": "true",
            "columns[0][searchable]": "true",
            "columns[0][orderable]": "true",
            "columns[0][search][value]": "",
            "columns[0][search][regex]": "false",
            "columns[1][data]": "type",
            "columns[1][name]": "",
            "columns[1][searchable]": "false",
            "columns[1][orderable]": "true",
            "columns[1][search][value]": "",
            "columns[1][search][regex]": "false",
            "columns[2][data]": "avatar",
            "columns[2][name]": "",
            "columns[2][searchable]": "false",
            "columns[2][orderable]": "false",
            "columns[2][search][value]": "",
            "columns[2][search][regex]": "false",
            "columns[3][data]": "owner",
            "columns[3][name]" :"",
            "columns[3][searchable]": "false",
            "columns[3][orderable]": "false",
            "columns[3][search][value]": "",
            "columns[3][search][regex]": "false",
            "columns[4][data]": "state",
            "columns[4][name]": "",
            "columns[4][searchable]": "false",
            "columns[4][orderable]": "true",
            "columns[4][search][value]": "",
            "columns[4][search][regex]": "false",
            "columns[5][data]": "",
            "columns[5][name]": "",
            "columns[5][searchable]": "false",
            "columns[5][orderable]": "false",
            "columns[5][search][value]": "",
            "columns[5][search][regex]": "false",
            "order[0][column]": "4",
            "order[0][dir]": "desc",
            "start": "0",
            "length": "2000",
            "search[value]": "",
            "search[regex]": "false"
        })

        numbers = {}
        j = json.loads(response.content)
        for no in j["data"]:
            n = FonialNumber(str(no["DT_RowId"]), str(no["number"][4:]), no["type"], no["state"], bool(no["cancel_at"]), bool(len(no["targets"])))
            numbers[n.number] = n
        logging.debug("found {} numbers: {}".format(len(numbers), numbers.keys()))
        return numbers

    def activateNumber(self, n):
        if n.state:
            logging.debug("{} already activated".format(n))
            return  # already activated

        logging.info("activating phone number {}".format(n))
        if self.args.dry_run:
            return

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/activate".format(n.id))
        soup = BeautifulSoup(response.content, 'html.parser')
        logging.debug(soup.prettify())
        csrf = soup.find("input", {"id": "form__token"}).get("value")
        assert csrf is not None
        logging.info("found CSRF token {}".format(csrf))

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/update/activate".format(n.id), data={
            "form[_token]": csrf
        })
        logging.debug(response.text)
        # assert response.text == "\\/system\\/number\\/{}\\/update\\/activate\\/success".format(n.id)
        n.state = True # store the new active state w/o reloading the truth from server -> a little risk!

    def deactivateNumber(self, n):
        if not n.state:
            logging.debug("{} already deactivated".format(n))
            return  # already deactivated

        if n.cancelled:
            logging.debug("{} already cancelled".format(n))
            return  # already cancelled

        logging.info("deactivating phone number {}".format(n))
        if self.args.dry_run:
            return

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/deactivate".format(n.id))
        soup = BeautifulSoup(response.content, 'html.parser')
        logging.debug(soup.prettify())
        csrf = soup.find("input", {"id": "form__token"}).get("value")
        assert csrf is not None
        logging.debug("found CSRF token {}".format(csrf))

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/update/deactivate".format(n.id), data={
            "form[_token]": csrf
        })
        logging.debug(response.text)
        n.state = False # store the new inactive state w/o reloading the truth from server -> a little risk!
        n.cancelled = False # store the new cancelled state w/o reloading the truth from server -> a little risk!

    def bind_device_to_number(self, n, d):
        if not n.state:
            logging.warning("can not bind device to the inactive number {}".format(n))
            return

        logging.info("binding device {} to phone number {}".format(d, n.number))
        if self.args.dry_run:
            return

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/edit/number".format(n.id))
        soup = BeautifulSoup(response.content, 'html.parser')
        csrf = soup.find("input", {"id": "fonial_databundle_number__token"}).get("value")
        assert csrf is not None
        logging.debug("found CSRF token {}".format(csrf))

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/update".format(n.id), {
            "fonial_databundle_number[settings][0][name]": " Regel 1",
            "fonial_databundle_number[settings][0][priority]": "0",
            "fonial_databundle_number[settings][0][active]": "1",
            "fonial_databundle_number[settings][0][redirectType]": "NORULE",
            "fonial_databundle_number[settings][0][redirectDaysFrom]": "",
            "fonial_databundle_number[settings][0][redirectDaysTo]": "",
            "fonial_databundle_number[settings][0][filterNumbersData]": "",
            "fonial_databundle_number[settings][0][forwardPromptData]": "",
            "fonial_databundle_number[settings][0][forwardPinData]": "",
            "fonial_databundle_number[settings][0][redirectSimultaneousPeriod]": "0",
            "fonial_databundle_number[settings][0][targets][0][target]": d,
            "fonial_databundle_number[settings][0][targets][0][delay]": "0",
            "fonial_databundle_number[base_target]": "",
            "fonial_databundle_number[type]": "VOICE",
            "fonial_databundle_number[_token]": csrf
        })
        logging.debug(response.text)

    def switch_number(self, source, destination, csrf=None):
        logging.info("switching {} to {}".format(source, destination))
        assert source.state is True
        assert destination.state is False

        if self.args.dry_run:
            return

        # get token
        if not csrf:
            response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/switch/number".format(source.id))
            soup = BeautifulSoup(response.content, 'html.parser')
            csrf = soup.find("input", {"id": "fonial_frontend_system_number__token"}).get("value")

        response = self.session.post("https://kundenkonto.fonial.de/system/number/{}/switch/number".format(source.id), {
            "fonial_frontend_system_number[unassign]": source.id,
            "fonial_frontend_system_number[assign]": destination.id,
            "fonial_frontend_system_number[_token]": csrf
        })

        logging.debug(response.text)

        if response.text != "{}":
            logging.warning("FAILED SWITCHING {} -> {}".format(source, destination))

        return csrf

    def switch_numbers(self):
        # TODO change to xlsx format
        with open(args.file, 'r', encoding='utf-8-sig', newline='') as src:
            reader = csv.reader(src, delimiter=';', quotechar='"')
            next(reader)  # skip header row

            mapping = {row[5]: row[1] for row in reader}
            logging.info("switching {} numbers".format(len(mapping)))
            assert len(mapping) > 0

            numbers = self.loadNumbers()
            csrf = None

            for a, b in mapping.items():
                csrf = self.switch_number(numbers[a], numbers[b], csrf)

    def  verify_user(self, user):
        logging.debug("verifying user {}".format(user))

        if user.fonial_device is None:
            logging.info("{}: no device configured".format(user.name))
            return

        response = self.session.get("https://kundenkonto.fonial.de/system/device/{}/ipdevice/edit".format(user.fonial_device.id))
        soup = BeautifulSoup(response.content, 'html.parser')
        # logging.debug(soup.prettify())
        assert soup.form.get("name") == "fonial_frontend_device_ip_device"

        errors = []
        target_name = soup.find("input", id="fonial_frontend_device_ip_device_targetName").get("value")
        if target_name != user.name:
            errors.append("target_name")
            logging.warning("{}: wrong device name {}".format(user.name, target_name))

        mac_address = soup.find("input", id="fonial_frontend_device_ip_device_mac").get("value")
        if mac_address != user.mac_address:
            errors.append("mac_address")
            logging.warning("{}: wrong device mac address {}. {} expected".format(user.name, mac_address, user.mac_address))

        num_list = soup.find("select", id="fonial_frontend_device_ip_device_outboundnum")
        outbound_num = num_list.find("option", selected=True).string
        if outbound_num[-9:] != user.number:
            errors.append("outbound_num")
            logging.warning("{}: wrong outbound {}. {} expected".format(user.name, outbound_num, user.number))

        extension = int(soup.find("input", id="fonial_frontend_device_ip_device_internalext").get("value"))
        if extension != user.extension:
            errors.append("extension")
            logging.warning("{}: wrong extension {}. {} expected".format(user.name, extension, user.extension))

        csrf = soup.find("input", id="fonial_frontend_device_ip_device__token").get("value")
        assert csrf is not None
        logging.debug("found CSRF token {}".format(csrf))

        if len(errors) == 0:
            logging.info("{}: all fine".format(user.name))
        else:
            logging.warning("{}: found errors {}".format(user.name, errors))

        # fix outbound num issue
        if len(errors) == 1 and errors[0] == "outbound_num":
            self.bind_number_to_device(user.name, user.fonial_device, user.fonial_number, csrf)


    def check_number_mapping(self, numbers, devices):
        wb = load_workbook(self.args.file, data_only=True)
        ws = wb.active
        # for i, row in enumerate(ws.values):
        for i, row in zip(range(200), ws.values):
            if i < 1:
                continue  # skip header row
            name = row[7]
            number = str(row[8])
            extension = row[9]
            mac_address = row[10]
            n = numbers[number] if number in numbers else None
            d = devices[mac_address] if mac_address is not None and mac_address in devices else None
            user = Employee(name, number, extension, mac_address, n, d)
            f.verify_user(user)

    def deactivate_unused_numbers(self):
        numbers = self.loadNumbers()
        for n in numbers.values():
            if n.state and not n.assigned and not n.cancelled:
                f.deactivateNumber(n)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Fonial admin batch tool')
    parser.add_argument('-a', '--account', required=True,
                        help='the fonial account id (usually 6 digit id)')
    parser.add_argument('-u', '--user', required=True,
                        help='the fonial backend username (usually an email address)')
    parser.add_argument('-p', '--password', required=False,
                        help='the fonial backend users password')
    parser.add_argument('-e', '--export', required=False,
                        help='export the current data in fonial into the specified file')
    parser.add_argument('-c', '--clean', required=False,
                        help='deactivate unused numbers')
    parser.add_argument('-s', '--switch-numbers', required=False,
                        help='switch current numbers if possible - previous numbers get invalid immediately')
    parser.add_argument('-d', '--debug', action="store_true", default=False,
                        help='set debug mode')
    parser.add_argument('-n', '--dry-run', action="store_true", default=False,
                        help='perform a trial run with no changes made')
    parser.add_argument('file',
                        help='the fonial excel file')
    args = parser.parse_args()

    if not args.password:
        args.password = getpass("password for {}".format(args.user))

    f = Fonial(args)

    if args.export:
        f.export()
    else:
        f.sync()

    # numbers = f.loadNumbers()
    # devices = f.loadDevices()

    # wb = load_workbook(args.file, data_only=True)
    # ws = wb.active
    # for i, row in enumerate(ws.values):
    # # for i, row in zip(range(2), ws.values):
    #     if i < 1:
    #         continue  # skip header row
    #     if row[13] == 'delete':
    #         name = row[7]
    #         number = str(row[8])
    #         mac_address = row[10]
    #         d = devices[mac_address] if mac_address is not None and mac_address in devices else None
    #         n = numbers[number] if number in numbers else None
    #         if d:
    #             f.delete_device(d)
    #         if n:
    #             f.deactivateNumber(n)

    # f.check_number_mapping(numbers, devices)
    # f.newDevices()
